import openpyxl
import pandas as pd


def read_configuration_data_from_excel(file, sheet_name="configuration"):
    df = pd.read_excel(file, sheet_name)
    return df.to_dict(orient="records")[0]


def read_test_data_from_excel(file, sheet_name, table_name):
    df = pd.read_excel(file, sheet_name, table_name)
    print(df.to_dict(orient="records")[0])
    return df.to_dict(orient="records")[0]


def read_data_from_excel_by_row(file, sheet_name, row_value):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    data = []
    for index, row in enumerate(sheet.rows):
        if row[0].value == row_value:
            for cell in row:
                data.append(cell.value)
    return tuple(data[1:])


# update a single excel column value
def update_excel_column_value(file, sheet_name, column_name, new_value):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows():
        if row[0].value == column_name:
            row[1].value = new_value
    workbook.save(file)
